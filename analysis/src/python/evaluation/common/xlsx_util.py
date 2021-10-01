import logging.config
from pathlib import Path
from typing import Union

import pandas as pd
from openpyxl import load_workbook, Workbook
from src.python.review.application_config import LanguageVersion

from analysis.src.python.evaluation.common.csv_util import ColumnName

logger = logging.getLogger(__name__)


def remove_sheet(workbook_path: Union[str, Path], sheet_name: str, to_raise_error: bool = False) -> None:
    try:
        workbook = load_workbook(workbook_path)
        workbook.remove(workbook[sheet_name])
        workbook.save(workbook_path)

    except KeyError as e:
        message = f'Sheet with specified name: {sheet_name} does not exist.'
        if to_raise_error:
            logger.exception(message)
            raise e
        else:
            logger.info(message)


def create_workbook(output_file_path: Path) -> Workbook:
    workbook = Workbook()
    workbook.save(output_file_path)
    return workbook


def write_dataframe_to_xlsx_sheet(xlsx_file_path: Union[str, Path], df: pd.DataFrame, sheet_name: str,
                                  mode: str = 'a', to_write_row_names: bool = False) -> None:
    """
    mode: str Available values are {'w', 'a'}. File mode to use (write or append).
    to_write_row_names: bool Write row names.
    """

    with pd.ExcelWriter(xlsx_file_path, mode=mode) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=to_write_row_names)


script_structure_rule = ('Please, make sure your XLSX-file matches following script standards: \n'
                         '1. Your XLSX-file or CSV-file should have 2 obligatory columns named:'
                         f'"{ColumnName.CODE.value}" & "{ColumnName.LANG.value}". \n'
                         f'"{ColumnName.CODE.value}" column -- relates to the code-sample. \n'
                         f'"{ColumnName.LANG.value}" column -- relates to the language of a '
                         'particular code-sample. \n'
                         '2. Your code samples should belong to the one of the supported languages. \n'
                         'Supported languages are: Java, Kotlin, Python. \n'
                         f'3. Check that "{ColumnName.LANG.value}" column cells are filled with '
                         'acceptable language-names: \n'
                         f'Acceptable language-names are: {LanguageVersion.PYTHON_3.value}, '
                         f'{LanguageVersion.JAVA_8.value} ,'
                         f'{LanguageVersion.JAVA_11.value} and {LanguageVersion.KOTLIN.value}.')